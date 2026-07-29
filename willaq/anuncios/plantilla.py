"""
Genera la plantilla en Excel donde el docente redacta los anuncios del curso.

Importante: este módulo NO publica nada en Blackboard, solo crea el archivo
Excel para que el profesor lo revise y edite con calma. La publicación
automática de anuncios es una función futura (ver willaq/anuncios/publicar.py).

'generar_plantilla()' es la función "pura": crea el archivo y devuelve un
diccionario con el resultado, sin usar print() ni input(). Así la puede
usar tanto el CLI como el panel web. 'generar_plantilla_cli()' es la
versión que sí interactúa por terminal (pregunta antes de sobrescribir).
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from willaq.config import DIR_PLANTILLAS

NOMBRE_ARCHIVO = "plantilla_anuncios.xlsx"

# Columnas de la plantilla: cada una es (encabezado, ancho de columna en Excel).
# Para agregar un campo nuevo en el futuro (por ejemplo "Curso" o "Adjunto"),
# basta con añadir una tupla aquí y el valor correspondiente en FILAS_EJEMPLO.
COLUMNAS = [
    ("Fecha", 14),
    ("Título", 35),
    ("Cuerpo", 70),
]

# Filas de ejemplo para que el profesor entienda cómo llenar cada columna.
# La fecha se deja como texto libre (dd/mm/aaaa) porque todavía no sabemos
# el formato exacto que exige el formulario de anuncios de Blackboard.
FILAS_EJEMPLO = [
    ("01/03/2026", "Bienvenida al curso", "Estimados alumnos, les damos la bienvenida al ciclo. Revisen el sílabo publicado en el curso."),
    ("15/03/2026", "Recordatorio de entrega", "Recuerden que la Práctica 1 vence el viernes a las 23:59. Cualquier duda, escríbanme por el foro."),
]

COLOR_ENCABEZADO = "4472C4"


def _configurar_hoja(hoja):
    """Escribe encabezados con estilo, ajusta anchos y agrega las filas de ejemplo."""

    for indice_columna, (encabezado, ancho) in enumerate(COLUMNAS, start=1):
        celda = hoja.cell(row=1, column=indice_columna, value=encabezado)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=COLOR_ENCABEZADO)
        celda.alignment = Alignment(horizontal="center", vertical="center")
        hoja.column_dimensions[celda.column_letter].width = ancho

    # Deja el encabezado siempre visible al desplazarse hacia abajo.
    hoja.freeze_panes = "A2"

    for indice_fila, fila in enumerate(FILAS_EJEMPLO, start=2):
        for indice_columna, valor in enumerate(fila, start=1):
            celda = hoja.cell(row=indice_fila, column=indice_columna, value=valor)
            celda.alignment = Alignment(wrap_text=True, vertical="top")


def _confirmar_sobrescritura(ruta_archivo) -> bool:
    """Pregunta al profesor antes de reemplazar un archivo que ya existe."""
    respuesta = input(
        f"El archivo '{ruta_archivo.name}' ya existe. ¿Deseas reemplazarlo? "
        "Se perderá lo que hayas escrito en él (s/n): "
    )
    return respuesta.strip().lower() in ("s", "si", "sí")


def generar_plantilla(forzar: bool = False) -> dict:
    """Crea el archivo Excel de anuncios. Función pura, sin print() ni input().

    Devuelve un diccionario con el resultado:
    - {"estado": "existe", "ruta": ...} si el archivo ya existe y 'forzar' es False.
    - {"estado": "ok", "ruta": ...} si el archivo se generó (o reemplazó) correctamente.
    """
    DIR_PLANTILLAS.mkdir(parents=True, exist_ok=True)
    ruta_archivo = DIR_PLANTILLAS / NOMBRE_ARCHIVO

    if ruta_archivo.exists() and not forzar:
        return {"estado": "existe", "ruta": str(ruta_archivo)}

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Anuncios"
    _configurar_hoja(hoja)
    libro.save(ruta_archivo)

    return {"estado": "ok", "ruta": str(ruta_archivo)}


def generar_plantilla_cli():
    """Versión para la terminal: pregunta antes de sobrescribir y muestra mensajes."""

    resultado = generar_plantilla()

    if resultado["estado"] == "existe":
        ruta_archivo = DIR_PLANTILLAS / NOMBRE_ARCHIVO
        if not _confirmar_sobrescritura(ruta_archivo):
            print("Operación cancelada. No se modificó el archivo existente.")
            return
        resultado = generar_plantilla(forzar=True)

    print("=" * 70)
    print("[OK] Plantilla de anuncios generada correctamente.")
    print(f"     Archivo: {resultado['ruta']}")
    print()
    print("Ábrelo con Excel, borra o edita las filas de ejemplo, y agrega una")
    print("fila por cada anuncio que quieras publicar (fecha, título y cuerpo).")
    print("=" * 70)
